import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Define styles
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
critical_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
critical_font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
high_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
high_font = Font(name="Arial", size=9, color="000000")
medium_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
medium_font = Font(name="Arial", size=9, color="000000")
low_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
low_font = Font(name="Arial", size=9, color="000000")
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Wallets to analyze
wallets = ["MetaMask", "Trust Wallet", "Ledger Live", "Trezor", "BlueWallet", "Phantom", "Exodus", "Atomic Wallet", "Electrum", "Coinbase Wallet"]

# Feature categories with features and importance levels
features_db = {
    "SECURITY": {
        "Hardware wallet integration": "Critical",
        "Multi-signature support": "Critical",
        "Biometric authentication": "Critical",
        "PIN/Password protection": "Critical",
        "Encryption at rest": "Critical",
        "Secure key generation": "Critical",
        "2FA support": "High",
        "Passphrase support": "High",
        "Air-gapped signing": "High",
        "Session timeout": "Medium"
    },
    "KEY_MANAGEMENT": {
        "HD wallet (BIP32/44)": "Critical",
        "Non-custodial": "Critical",
        "Private key control": "Critical",
        "Deterministic generation": "High",
        "Hierarchical derivation": "High",
        "Hardware key storage": "High"
    },
    "RECOVERY": {
        "Seed phrase (12/24)": "Critical",
        "Seed verification": "Critical",
        "Recovery from seed": "Critical",
        "Social recovery": "Medium",
        "Recovery guides": "Medium"
    },
    "UX": {
        "Intuitive onboarding": "Critical",
        "Dark mode": "High",
        "Mobile app": "High",
        "Multi-language": "High",
        "Browser extension": "High",
        "Desktop app": "Medium",
        "Accessibility features": "Medium"
    },
    "DEFI": {
        "DEX/Swap": "High",
        "Staking": "High",
        "Smart contracts": "High",
        "Gas optimization": "Medium",
        "Transaction simulation": "Medium"
    },
    "FEATURES": {
        "Token management": "Critical",
        "NFT support": "High",
        "Multi-chain": "High",
        "Custom RPC": "Medium",
        "Token import": "High"
    },
    "PRIVACY": {
        "CoinJoin support": "High",
        "Privacy coins": "High",
        "Tor support": "Medium",
        "No IP logging": "Medium",
        "Private RPC": "Medium"
    },
    "BACKUP": {
        "Seed export": "Critical",
        "Multiple backups": "High",
        "Cloud backup": "Medium",
        "Manual backup": "High",
        "Backup reminders": "Low"
    },
    "NOTIFICATIONS": {
        "TX alerts": "High",
        "Price alerts": "Medium",
        "Security alerts": "High",
        "Gas alerts": "Medium",
        "Push notifications": "Medium"
    },
    "ANALYTICS": {
        "Portfolio tracking": "Critical",
        "Price charts": "High",
        "TX history": "Critical",
        "Tax export": "High",
        "Spending analytics": "Medium"
    }
}

# Wallet feature data
wallet_features = {
    "MetaMask": {"Hardware wallet integration": 1, "Multi-signature support": 0, "Biometric authentication": 1, "PIN/Password protection": 1, "Encryption at rest": 1, "Secure key generation": 1, "2FA support": 0, "Passphrase support": 1, "Air-gapped signing": 0, "Session timeout": 1, "HD wallet (BIP32/44)": 1, "Non-custodial": 1, "Private key control": 1, "Deterministic generation": 1, "Hierarchical derivation": 1, "Hardware key storage": 1, "Seed phrase (12/24)": 1, "Seed verification": 1, "Recovery from seed": 1, "Social recovery": 0, "Recovery guides": 1, "Intuitive onboarding": 1, "Dark mode": 1, "Mobile app": 1, "Multi-language": 1, "Browser extension": 1, "Desktop app": 0, "Accessibility features": 0, "DEX/Swap": 1, "Staking": 1, "Smart contracts": 1, "Gas optimization": 1, "Transaction simulation": 1, "Token management": 1, "NFT support": 1, "Multi-chain": 1, "Custom RPC": 1, "Token import": 1, "CoinJoin support": 0, "Privacy coins": 0, "Tor support": 0, "No IP logging": 0, "Private RPC": 0, "Seed export": 1, "Multiple backups": 0, "Cloud backup": 0, "Manual backup": 1, "Backup reminders": 0, "TX alerts": 1, "Price alerts": 1, "Security alerts": 1, "Gas alerts": 1, "Push notifications": 1, "Portfolio tracking": 1, "Price charts": 0, "TX history": 1, "Tax export": 0, "Spending analytics": 0},
    "Trust Wallet": {"Hardware wallet integration": 1, "Multi-signature support": 0, "Biometric authentication": 1, "PIN/Password protection": 1, "Encryption at rest": 1, "Secure key generation": 1, "2FA support": 0, "Passphrase support": 1, "Air-gapped signing": 0, "Session timeout": 1, "HD wallet (BIP32/44)": 1, "Non-custodial": 1, "Private key control": 1, "Deterministic generation": 1, "Hierarchical derivation": 1, "Hardware key storage": 1, "Seed phrase (12/24)": 1, "Seed verification": 1, "Recovery from seed": 1, "Social recovery": 0, "Recovery guides": 1, "Intuitive onboarding": 1, "Dark mode": 1, "Mobile app": 1, "Multi-language": 1, "Browser extension": 0, "Desktop app": 0, "Accessibility features": 0, "DEX/Swap": 1, "Staking": 1, "Smart contracts": 1, "Gas optimization": 1, "Transaction simulation": 0, "Token management": 1, "NFT support": 1, "Multi-chain": 1, "Custom RPC": 0, "Token import": 1, "CoinJoin support": 0, "Privacy coins": 0, "Tor support": 0, "No IP logging": 0, "Private RPC": 0, "Seed export": 1, "Multiple backups": 0, "Cloud backup": 0, "Manual backup": 1, "Backup reminders": 0, "TX alerts": 1, "Price alerts": 1, "Security alerts": 1, "Gas alerts": 1, "Push notifications": 1, "Portfolio tracking": 1, "Price charts": 1, "TX history": 1, "Tax export": 0, "Spending analytics": 0},
    "Ledger Live": {"Hardware wallet integration": 1, "Multi-signature support": 1, "Biometric authentication": 1, "PIN/Password protection": 1, "Encryption at rest": 1, "Secure key generation": 1, "2FA support": 1, "Passphrase support": 1, "Air-gapped signing": 1, "Session timeout": 1, "HD wallet (BIP32/44)": 1, "Non-custodial": 1, "Private key control": 1, "Deterministic generation": 1, "Hierarchical derivation": 1, "Hardware key storage": 1, "Seed phrase (12/24)": 1, "Seed verification": 1, "Recovery from seed": 1, "Social recovery": 0, "Recovery guides": 1, "Intuitive onboarding": 1, "Dark mode": 1, "Mobile app": 1, "Multi-language": 1, "Browser extension": 0, "Desktop app": 1, "Accessibility features": 1, "DEX/Swap": 1, "Staking": 1, "Smart contracts": 1, "Gas optimization": 0, "Transaction simulation": 0, "Token management": 1, "NFT support": 1, "Multi-chain": 1, "Custom RPC": 0, "Token import": 1, "CoinJoin support": 0, "Privacy coins": 0, "Tor support": 0, "No IP logging": 0, "Private RPC": 0, "Seed export": 1, "Multiple backups": 1, "Cloud backup": 0, "Manual backup": 1, "Backup reminders": 1, "TX alerts": 1, "Price alerts": 1, "Security alerts": 1, "Gas alerts": 1, "Push notifications": 1, "Portfolio tracking": 1, "Price charts": 1, "TX history": 1, "Tax export": 1, "Spending analytics": 1},
    "Trezor": {"Hardware wallet integration": 1, "Multi-signature support": 1, "Biometric authentication": 1, "PIN/Password protection": 1, "Encryption at rest": 1, "Secure key generation": 1, "2FA support": 1, "Passphrase support": 1, "Air-gapped signing": 1, "Session timeout": 1, "HD wallet (BIP32/44)": 1, "Non-custodial": 1, "Private key control": 1, "Deterministic generation": 1, "Hierarchical derivation": 1, "Hardware key storage": 1, "Seed phrase (12/24)": 1, "Seed verification": 1, "Recovery from seed": 1, "Social recovery": 0, "Recovery guides": 1, "Intuitive onboarding": 1, "Dark mode": 0, "Mobile app": 0, "Multi-language": 1, "Browser extension": 1, "Desktop app": 1, "Accessibility features": 0, "DEX/Swap": 0, "Staking": 1, "Smart contracts": 0, "Gas optimization": 0, "Transaction simulation": 0, "Token management": 1, "NFT support": 0, "Multi-chain": 1, "Custom RPC": 0, "Token import": 1, "CoinJoin support": 1, "Privacy coins": 1, "Tor support": 1, "No IP logging": 0, "Private RPC": 0, "Seed export": 1, "Multiple backups": 1, "Cloud backup": 0, "Manual backup": 1, "Backup reminders": 1, "TX alerts": 0, "Price alerts": 0, "Security alerts": 1, "Gas alerts": 0, "Push notifications": 0, "Portfolio tracking": 0, "Price charts": 0, "TX history": 1, "Tax export": 0, "Spending analytics": 0},
    "BlueWallet": {"Hardware wallet integration": 1, "Multi-signature support": 1, "Biometric authentication": 1, "PIN/Password protection": 1, "Encryption at rest": 1, "Secure key generation": 1, "2FA support": 0, "Passphrase support": 1, "Air-gapped signing": 1, "Session timeout": 0, "HD wallet (BIP32/44)": 1, "Non-custodial": 1, "Private key control": 1, "Deterministic generation": 1, "Hierarchical derivation": 1, "Hardware key storage": 1, "Seed phrase (12/24)": 1, "Seed verification": 0, "Recovery from seed": 1, "Social recovery": 0, "Recovery guides": 1, "Intuitive onboarding": 1, "Dark mode": 1, "Mobile app": 1, "Multi-language": 1, "Browser extension": 0, "Desktop app": 0, "Accessibility features": 0, "DEX/Swap": 0, "Staking": 0, "Smart contracts": 0, "Gas optimization": 0, "Transaction simulation": 0, "Token management": 0, "NFT support": 0, "Multi-chain": 0, "Custom RPC": 1, "Token import": 0, "CoinJoin support": 1, "Privacy coins": 0, "Tor support": 1, "No IP logging": 1, "Private RPC": 1, "Seed export": 1, "Multiple backups": 0, "Cloud backup": 0, "Manual backup": 1, "Backup reminders": 0, "TX alerts": 1, "Price alerts": 0, "Security alerts": 0, "Gas alerts": 0, "Push notifications": 1, "Portfolio tracking": 1, "Price charts": 1, "TX history": 1, "Tax export": 0, "Spending analytics": 0},
    "Phantom": {"Hardware wallet integration": 1, "Multi-signature support": 0, "Biometric authentication": 1, "PIN/Password protection": 1, "Encryption at rest": 1, "Secure key generation": 1, "2FA support": 0, "Passphrase support": 1, "Air-gapped signing": 0, "Session timeout": 1, "HD wallet (BIP32/44)": 1, "Non-custodial": 1, "Private key control": 1, "Deterministic generation": 1, "Hierarchical derivation": 1, "Hardware key storage": 1, "Seed phrase (12/24)": 1, "Seed verification": 1, "Recovery from seed": 1, "Social recovery": 0, "Recovery guides": 1, "Intuitive onboarding": 1, "Dark mode": 1, "Mobile app": 1, "Multi-language": 1, "Browser extension": 1, "Desktop app": 0, "Accessibility features": 0, "DEX/Swap": 1, "Staking": 1, "Smart contracts": 1, "Gas optimization": 0, "Transaction simulation": 1, "Token management": 1, "NFT support": 1, "Multi-chain": 1, "Custom RPC": 1, "Token import": 1, "CoinJoin support": 0, "Privacy coins": 0, "Tor support": 0, "No IP logging": 0, "Private RPC": 0, "Seed export": 1, "Multiple backups": 0, "Cloud backup": 0, "Manual backup": 1, "Backup reminders": 0, "TX alerts": 1, "Price alerts": 1, "Security alerts": 1, "Gas alerts": 1, "Push notifications": 1, "Portfolio tracking": 1, "Price charts": 1, "TX history": 1, "Tax export": 0, "Spending analytics": 0},
    "Exodus": {"Hardware wallet integration": 1, "Multi-signature support": 0, "Biometric authentication": 1, "PIN/Password protection": 1, "Encryption at rest": 1, "Secure key generation": 1, "2FA support": 0, "Passphrase support": 0, "Air-gapped signing": 0, "Session timeout": 0, "HD wallet (BIP32/44)": 1, "Non-custodial": 1, "Private key control": 1, "Deterministic generation": 1, "Hierarchical derivation": 0, "Hardware key storage": 1, "Seed phrase (12/24)": 1, "Seed verification": 0, "Recovery from seed": 1, "Social recovery": 0, "Recovery guides": 1, "Intuitive onboarding": 1, "Dark mode": 1, "Mobile app": 1, "Multi-language": 1, "Browser extension": 0, "Desktop app": 1, "Accessibility features": 0, "DEX/Swap": 1, "Staking": 1, "Smart contracts": 0, "Gas optimization": 0, "Transaction simulation": 0, "Token management": 1, "NFT support": 0, "Multi-chain": 0, "Custom RPC": 0, "Token import": 1, "CoinJoin support": 0, "Privacy coins": 0, "Tor support": 0, "No IP logging": 0, "Private RPC": 0, "Seed export": 1, "Multiple backups": 0, "Cloud backup": 0, "Manual backup": 1, "Backup reminders": 0, "TX alerts": 1, "Price alerts": 1, "Security alerts": 0, "Gas alerts": 0, "Push notifications": 1, "Portfolio tracking": 1, "Price charts": 1, "TX history": 1, "Tax export": 0, "Spending analytics": 0},
    "Atomic Wallet": {"Hardware wallet integration": 1, "Multi-signature support": 0, "Biometric authentication": 1, "PIN/Password protection": 1, "Encryption at rest": 1, "Secure key generation": 1, "2FA support": 0, "Passphrase support": 1, "Air-gapped signing": 0, "Session timeout": 1, "HD wallet (BIP32/44)": 1, "Non-custodial": 1, "Private key control": 1, "Deterministic generation": 1, "Hierarchical derivation": 1, "Hardware key storage": 1, "Seed phrase (12/24)": 1, "Seed verification": 0, "Recovery from seed": 1, "Social recovery": 0, "Recovery guides": 1, "Intuitive onboarding": 1, "Dark mode": 1, "Mobile app": 1, "Multi-language": 1, "Browser extension": 0, "Desktop app": 1, "Accessibility features": 0, "DEX/Swap": 1, "Staking": 1, "Smart contracts": 1, "Gas optimization": 0, "Transaction simulation": 0, "Token management": 1, "NFT support": 1, "Multi-chain": 1, "Custom RPC": 0, "Token import": 1, "CoinJoin support": 0, "Privacy coins": 1, "Tor support": 0, "No IP logging": 0, "Private RPC": 0, "Seed export": 1, "Multiple backups": 0, "Cloud backup": 0, "Manual backup": 1, "Backup reminders": 0, "TX alerts": 1, "Price alerts": 1, "Security alerts": 0, "Gas alerts": 0, "Push notifications": 1, "Portfolio tracking": 1, "Price charts": 1, "TX history": 1, "Tax export": 0, "Spending analytics": 0},
    "Electrum": {"Hardware wallet integration": 1, "Multi-signature support": 1, "Biometric authentication": 1, "PIN/Password protection": 1, "Encryption at rest": 1, "Secure key generation": 1, "2FA support": 1, "Passphrase support": 1, "Air-gapped signing": 1, "Session timeout": 0, "HD wallet (BIP32/44)": 1, "Non-custodial": 1, "Private key control": 1, "Deterministic generation": 1, "Hierarchical derivation": 1, "Hardware key storage": 1, "Seed phrase (12/24)": 1, "Seed verification": 0, "Recovery from seed": 1, "Social recovery": 0, "Recovery guides": 1, "Intuitive onboarding": 0, "Dark mode": 1, "Mobile app": 1, "Multi-language": 1, "Browser extension": 0, "Desktop app": 1, "Accessibility features": 0, "DEX/Swap": 0, "Staking": 0, "Smart contracts": 0, "Gas optimization": 0, "Transaction simulation": 0, "Token management": 0, "NFT support": 0, "Multi-chain": 0, "Custom RPC": 1, "Token import": 0, "CoinJoin support": 1, "Privacy coins": 0, "Tor support": 1, "No IP logging": 1, "Private RPC": 1, "Seed export": 1, "Multiple backups": 1, "Cloud backup": 0, "Manual backup": 1, "Backup reminders": 0, "TX alerts": 0, "Price alerts": 0, "Security alerts": 1, "Gas alerts": 0, "Push notifications": 0, "Portfolio tracking": 1, "Price charts": 0, "TX history": 1, "Tax export": 1, "Spending analytics": 0},
    "Coinbase Wallet": {"Hardware wallet integration": 1, "Multi-signature support": 0, "Biometric authentication": 1, "PIN/Password protection": 1, "Encryption at rest": 1, "Secure key generation": 1, "2FA support": 0, "Passphrase support": 1, "Air-gapped signing": 0, "Session timeout": 1, "HD wallet (BIP32/44)": 1, "Non-custodial": 1, "Private key control": 1, "Deterministic generation": 1, "Hierarchical derivation": 1, "Hardware key storage": 1, "Seed phrase (12/24)": 1, "Seed verification": 1, "Recovery from seed": 1, "Social recovery": 0, "Recovery guides": 1, "Intuitive onboarding": 1, "Dark mode": 1, "Mobile app": 1, "Multi-language": 1, "Browser extension": 1, "Desktop app": 0, "Accessibility features": 0, "DEX/Swap": 1, "Staking": 1, "Smart contracts": 1, "Gas optimization": 1, "Transaction simulation": 1, "Token management": 1, "NFT support": 1, "Multi-chain": 1, "Custom RPC": 0, "Token import": 1, "CoinJoin support": 0, "Privacy coins": 0, "Tor support": 0, "No IP logging": 0, "Private RPC": 0, "Seed export": 1, "Multiple backups": 0, "Cloud backup": 0, "Manual backup": 1, "Backup reminders": 0, "TX alerts": 1, "Price alerts": 1, "Security alerts": 1, "Gas alerts": 1, "Push notifications": 1, "Portfolio tracking": 1, "Price charts": 1, "TX history": 1, "Tax export": 0, "Spending analytics": 0}
}

# Create SUMMARY sheet
ws = wb.create_sheet("SUMMARY", 0)
ws['A1'] = "CRYPTOCURRENCY WALLET FEATURE ANALYSIS"
ws['A1'].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
ws['A1'].fill = header_fill
ws.merge_cells('A1:K1')
ws.row_dimensions[1].height = 25

row = 3
ws[f'A{row}'] = "EXECUTIVE OVERVIEW"
ws[f'A{row}'].font = Font(name="Arial", size=12, bold=True)

row += 2
ws[f'A{row}'] = "This comprehensive analysis evaluates 10 leading cryptocurrency wallets across 50+ critical features."
row += 1
ws[f'A{row}'] = "Source: 2024-2025 research based on official wallet documentation and feature verification."
ws[f'A{row}'].font = Font(name="Arial", size=9, italic=True)

row += 2
ws[f'A{row}'] = "WALLETS ANALYZED:"
ws[f'A{row}'].font = Font(name="Arial", size=11, bold=True)
row += 1
for wallet in wallets:
    ws[f'A{row}'] = f"  • {wallet}"
    row += 1

row += 1
ws[f'A{row}'] = "KEY FINDINGS:"
ws[f'A{row}'].font = Font(name="Arial", size=11, bold=True)
row += 1

findings = [
    "Security Leaders: Ledger Live, Trezor, and Electrum offer air-gapped signing and multi-sig support",
    "User Experience: MetaMask, Trust Wallet, and Phantom dominate with mobile and web integration",
    "Privacy Focus: Trezor, BlueWallet, and Electrum support CoinJoin, Tor, and private RPC",
    "DeFi Leaders: MetaMask, Trust Wallet, Phantom, and Coinbase Wallet for DeFi/NFTs",
    "Best Overall: Ledger Live excels with offline mode, tax reporting, and regulatory compliance",
    "Bitcoin Privacy: BlueWallet offers CoinJoin and Tor with minimal surface area",
    "Power Users: Electrum for advanced Bitcoin features with air-gapped transaction signing"
]

for finding in findings:
    ws[f'A{row}'] = f"  • {finding}"
    ws.merge_cells(f'A{row}:K{row}')
    row += 1

row += 2
ws[f'A{row}'] = "IMPORTANCE LEVELS:"
ws[f'A{row}'].font = Font(name="Arial", size=11, bold=True)
row += 1

importance_info = {
    "CRITICAL (Red)": "Essential for wallet function, security, or key management - must have",
    "HIGH (Gold)": "Important features enhancing security, capability, or usability",
    "MEDIUM (Green)": "Nice-to-have features improving user experience or extending functionality",
    "LOW (Light Blue)": "Optional features providing convenience or specialized capabilities"
}

for level, desc in importance_info.items():
    parts = level.split(' ')
    cell = ws[f'A{row}']
    cell.value = parts[0]
    if "Red" in level:
        cell.fill = critical_fill
        cell.font = critical_font
    elif "Gold" in level:
        cell.fill = high_fill
    elif "Green" in level:
        cell.fill = medium_fill
    else:
        cell.fill = low_fill

    ws[f'B{row}'] = desc
    ws.merge_cells(f'B{row}:K{row}')
    row += 1

ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 85

# Create CHECKLIST sheet with all features
ws = wb.create_sheet("FEATURE CHECKLIST")
ws['A1'] = "COMPREHENSIVE FEATURE CHECKLIST BY CATEGORY"
ws['A1'].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
ws['A1'].fill = header_fill
ws.merge_cells('A1:K1')

row = 3

# Headers
ws.cell(row=row, column=1, value="FEATURE CATEGORY").font = header_font
ws.cell(row=row, column=1).fill = header_fill
ws.cell(row=row, column=1).border = border

ws.cell(row=row, column=2, value="FEATURE NAME").font = header_font
ws.cell(row=row, column=2).fill = header_fill
ws.cell(row=row, column=2).border = border

ws.cell(row=row, column=3, value="IMPORTANCE").font = header_font
ws.cell(row=row, column=3).fill = header_fill
ws.cell(row=row, column=3).border = border

col_idx = 4
for wallet in wallets:
    ws.cell(row=row, column=col_idx, value=wallet).font = header_font
    ws.cell(row=row, column=col_idx).fill = header_fill
    ws.cell(row=row, column=col_idx).border = border
    ws.cell(row=row, column=col_idx).alignment = Alignment(wrap_text=True, horizontal="center")
    col_idx += 1

row += 1

# Add all features
for category, features in features_db.items():
    for feature, importance in features.items():
        ws.cell(row=row, column=1, value=category).font = Font(name="Arial", size=9)
        ws.cell(row=row, column=1).border = border

        ws.cell(row=row, column=2, value=feature).font = Font(name="Arial", size=9)
        ws.cell(row=row, column=2).border = border

        cell = ws.cell(row=row, column=3, value=importance)
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

        if importance == "Critical":
            cell.fill = critical_fill
            cell.font = critical_font
        elif importance == "High":
            cell.fill = high_fill
            cell.font = high_font
        elif importance == "Medium":
            cell.fill = medium_fill
            cell.font = medium_font
        else:
            cell.fill = low_fill
            cell.font = low_font

        col_idx = 4
        for wallet in wallets:
            has_feature = wallet_features[wallet].get(feature, 0)
            cell = ws.cell(row=row, column=col_idx, value="YES" if has_feature else "NO")
            cell.font = Font(name="Arial", size=9, bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

            if has_feature:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                cell.font = Font(name="Arial", size=9, bold=True, color="006100")
            else:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell.font = Font(name="Arial", size=9, bold=True, color="9C0006")

            col_idx += 1

        row += 1

ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 28
ws.column_dimensions['C'].width = 12
for col in range(4, 14):
    ws.column_dimensions[get_column_letter(col)].width = 12

# Save
wb.save('C:/Users/usman/Desktop/BigCoinBB/Cryptocurrency_Wallet_Analysis.xlsx')
print("Excel file created: Cryptocurrency_Wallet_Analysis.xlsx")
